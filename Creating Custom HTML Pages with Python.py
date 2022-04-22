#!/usr/bin/env python
# coding: utf-8

# In[21]:


pip install openpyxl


# In[18]:


import xlrd
book = xlrd.open_workbook("C:\seleniumdrivers\Data Science Jobs.x


# In[31]:


path = "C:\seleniumdrivers\Data Science Jobs.xls" 


# In[24]:


# Importing openpyxl module
import openpyxl


# In[25]:


# Location of the file
path = "C:\seleniumdrivers\Dat Science Jobs.xlsx" 


# In[26]:


# Workbook object is created to open the workbook
wb_obj = openpyxl.load_workbook(path)


# In[27]:


# Getting the workbook active sheet object and workbook object is created
sheet_obj = wb_obj.active


# In[28]:


# Cell object is created by the use of the sheet object method
cell_obj = sheet_obj['A1' : 'B15']


# In[29]:


# Printing the value of the cell object by using the value attribute
for cell1, cell2 in cell_obj:
    print(cell1.value, cell2.value)


# In[33]:


# Getting workbook active object from the active attribute
sheet_obj = wb_obj.active


# In[35]:


# Getting the value of maximum rows and columns
row = sheet_obj.max_row
column = sheet_obj.max_column


# In[36]:


print("\nValue of second column")
for i in range(2, row + 2):
    cell_obj = sheet_obj.cell(row = i, column = 2)
    print(cell_obj.value)


# In[37]:


print("\nValue of second column")
for i in range(3, row + 3):
    cell_obj = sheet_obj.cell(row = i, column = 3)
    print(cell_obj.value)


# In[38]:


print("\nValue of second column")
for i in range(4, row + 4):
    cell_obj = sheet_obj.cell(row = i, column = 4)
    print(cell_obj.value)


# In[ ]:




